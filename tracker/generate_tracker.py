"""
Daily IT Project Manager Excel Tracker Generator
Generates daily_project_tracker.xlsx with 7 fully formatted sheets.
Usage: python generate_tracker.py
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colour constants
# ---------------------------------------------------------------------------
DARK_BLUE = "1F3864"
WHITE = "FFFFFF"
LIGHT_BLUE = "DCE6F1"
GREEN_FILL = "C6EFCE"
RED_FILL = "FFC7CE"
AMBER_FILL = "FFEB9C"
ORANGE_FILL = "F4B942"

# Sheet tab colours
TAB_COLOURS = [
    "1F3864",  # DASHBOARD        – dark blue
    "2E75B6",  # SPRINT TRACKER   – medium blue
    "C00000",  # PRODUCTION ISSUES – red
    "7030A0",  # UAT TRACKER      – purple
    "FF6600",  # QA ISSUES        – orange
    "375623",  # PROD RELEASE     – dark green
    "595959",  # DAILY STANDUP    – grey
]

# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------

def _header_font():
    return Font(bold=True, color=WHITE, name="Calibri", size=11)


def _header_fill():
    return PatternFill("solid", fgColor=DARK_BLUE)


def _alt_fill(row_idx):
    """Return alternating row fill (1-based data row index)."""
    colour = LIGHT_BLUE if row_idx % 2 == 0 else WHITE
    return PatternFill("solid", fgColor=colour)


def _thin_border():
    thin = Side(border_style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_header(ws, headers, row=1):
    """Write a bold, dark-blue header row."""
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[row].height = 20


def _apply_data_row(ws, row_idx, values, excel_row):
    """Write a data row with alternating fill and border."""
    fill = _alt_fill(row_idx)
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=excel_row, column=col, value=value)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = _thin_border()


def _auto_size_columns(ws, min_width=15, max_width=40):
    """Auto-size columns based on content length."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _freeze_and_filter(ws, filter_ref=None):
    """Freeze top row and enable auto-filter."""
    ws.freeze_panes = "A2"
    if filter_ref:
        ws.auto_filter.ref = filter_ref


def _add_cf_status(ws, col_letter, data_rows, value, fill_colour):
    """Add conditional formatting rule for a specific status value in a column."""
    fill = PatternFill("solid", fgColor=fill_colour)
    ws.conditional_formatting.add(
        f"A2:{get_column_letter(ws.max_column or 20)}{data_rows + 1}",
        CellIsRule(operator="equal", formula=[f'"{value}"'], fill=fill)
    )


# ---------------------------------------------------------------------------
# Sheet 1: DASHBOARD
# ---------------------------------------------------------------------------

def create_dashboard(wb, sprint_sheet_title):
    ws = wb.create_sheet("DASHBOARD")

    # Title block
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "📊 Daily IT Project Manager Dashboard"
    title_cell.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    title_cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Sprint info labels
    sprint_info = [
        ("Sprint Name:", "Sprint 42"),
        ("Sprint Start Date:", "2026-05-25"),
        ("Sprint End Date:", "2026-06-07"),
    ]
    for i, (label, value) in enumerate(sprint_info, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, name="Calibri")
        ws.cell(row=i, column=2, value=value)

    # Separator
    ws.row_dimensions[5].height = 8

    # Summary metrics header
    ws.merge_cells("A6:D6")
    hdr = ws["A6"]
    hdr.value = "Sprint & Issue Summary"
    hdr.font = Font(bold=True, color=WHITE, name="Calibri")
    hdr.fill = PatternFill("solid", fgColor=DARK_BLUE)
    hdr.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 20

    # Metric rows: (label, formula, fill_hex)
    status_col = "F"  # Status column in SPRINT TRACKER
    metrics = [
        ("Total Tickets",
         f"=COUNTA('{sprint_sheet_title}'!A2:A1000)",
         "BDD7EE"),
        ("Done",
         f"=COUNTIF('{sprint_sheet_title}'!F:F,\"Done\")",
         GREEN_FILL),
        ("In Progress",
         f"=COUNTIF('{sprint_sheet_title}'!F:F,\"In Progress\")",
         AMBER_FILL),
        ("Blocked",
         f"=COUNTIF('{sprint_sheet_title}'!F:F,\"Blocked\")",
         RED_FILL),
        ("To Do",
         f"=COUNTIF('{sprint_sheet_title}'!F:F,\"To Do\")",
         "BDD7EE"),
        ("Open PROD Issues",
         "=COUNTIF('PRODUCTION ISSUES'!I:I,\"Open\")",
         RED_FILL),
        ("Open UAT Issues",
         "=COUNTIF('UAT TRACKER'!F:F,\"In Testing\")",
         AMBER_FILL),
        ("Open QA Issues",
         "=COUNTIF('QA ISSUES'!H:H,\"Open\")",
         RED_FILL),
    ]

    for idx, (label, formula, fill_hex) in enumerate(metrics, start=7):
        label_cell = ws.cell(row=idx, column=1, value=label)
        label_cell.font = Font(bold=True, name="Calibri")
        label_cell.border = _thin_border()

        value_cell = ws.cell(row=idx, column=2, value=formula)
        value_cell.fill = PatternFill("solid", fgColor=fill_hex)
        value_cell.font = Font(bold=True, name="Calibri")
        value_cell.alignment = Alignment(horizontal="center")
        value_cell.border = _thin_border()

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Sheet 2: SPRINT TRACKER
# ---------------------------------------------------------------------------

def create_sprint_tracker(wb):
    ws = wb.create_sheet("SPRINT TRACKER")

    headers = [
        "Ticket ID", "Summary", "Team", "Assignee", "Priority", "Status",
        "Story Points", "Sprint Day Created", "Due Date", "Actual Completion",
        "Blocker Notes", "Comments/Updates", "Going to QA?", "Going to UAT?",
        "Going to PROD?", "Traceability ID",
    ]
    _apply_header(ws, headers)

    data = [
        ("PROJ-101", "User authentication via SSO", "Backend", "Alice Johnson",
         "High", "Done", 5, "2026-05-25", "2026-05-28", "2026-05-28",
         "", "Completed ahead of schedule", "Yes", "Yes", "Yes", "REL-2026-05-31"),
        ("PROJ-102", "Dashboard widget rendering bug", "Frontend", "Bob Smith",
         "Critical", "In Progress", 3, "2026-05-25", "2026-05-27", "",
         "Waiting for design assets", "Blocked on Figma file from design team",
         "No", "No", "Yes", "REL-2026-05-31"),
        ("PROJ-103", "Payment gateway timeout handling", "Backend", "Carol White",
         "High", "Blocked", 8, "2026-05-26", "2026-05-30", "",
         "Third-party API rate limit issue", "Escalated to vendor – awaiting response",
         "Yes", "Yes", "Yes", "REL-2026-06-05"),
        ("PROJ-104", "Automated regression test suite update", "QA", "David Lee",
         "Medium", "In Progress", 5, "2026-05-26", "2026-05-29", "",
         "", "60% complete; focusing on payment module", "No", "No", "No", ""),
        ("PROJ-105", "Export to PDF feature", "Frontend", "Eve Martinez",
         "Low", "To Do", 3, "2026-05-27", "2026-06-02", "",
         "", "Not yet started; dependency on PROJ-101", "Yes", "Yes", "Yes",
         "REL-2026-06-05"),
    ]

    for row_idx, row_data in enumerate(data, start=1):
        _apply_data_row(ws, row_idx, row_data, row_idx + 1)

    # Conditional formatting on Status column (F)
    num_rows = len(data) + 1
    data_range = f"A2:P{num_rows + 1}"

    red_fill = PatternFill("solid", fgColor=RED_FILL)
    green_fill = PatternFill("solid", fgColor=GREEN_FILL)
    amber_fill = PatternFill("solid", fgColor=AMBER_FILL)

    ws.conditional_formatting.add(
        f"F2:F{num_rows + 1}",
        CellIsRule(operator="equal", formula=['"Blocked"'], fill=red_fill)
    )
    ws.conditional_formatting.add(
        f"F2:F{num_rows + 1}",
        CellIsRule(operator="equal", formula=['"Done"'], fill=green_fill)
    )
    ws.conditional_formatting.add(
        f"F2:F{num_rows + 1}",
        CellIsRule(operator="equal", formula=['"In Progress"'], fill=amber_fill)
    )

    _freeze_and_filter(ws, f"A1:P{num_rows + 1}")
    _auto_size_columns(ws)


# ---------------------------------------------------------------------------
# Sheet 3: PRODUCTION ISSUES
# ---------------------------------------------------------------------------

def create_production_issues(wb):
    ws = wb.create_sheet("PRODUCTION ISSUES")

    headers = [
        "Issue ID", "Date Reported", "Severity", "Environment", "Description",
        "Reported By", "Assigned To", "Root Cause", "Status", "Fix Ticket ID",
        "Resolution Date", "Postmortem Done?", "Comms Sent?",
    ]
    _apply_header(ws, headers)

    data = [
        ("INC-001", "2026-05-28", "P1", "PROD",
         "Payment service down – all transactions failing",
         "Customer Support", "Carol White",
         "DB connection pool exhausted due to missing index on orders table",
         "Resolved", "PROJ-110", "2026-05-28", "Yes", "Yes"),
        ("INC-002", "2026-05-29", "P2", "PROD",
         "Login page intermittently returning 502",
         "Internal Monitoring", "Alice Johnson",
         "Load balancer health check misconfiguration after deploy",
         "In Fix", "PROJ-111", "", "No", "Yes"),
        ("INC-003", "2026-05-30", "P3", "PROD",
         "PDF export formatting broken for Safari users",
         "QA Team", "Eve Martinez",
         "CSS print media query not applied in WebKit",
         "Open", "PROJ-105", "", "No", "No"),
    ]

    for row_idx, row_data in enumerate(data, start=1):
        _apply_data_row(ws, row_idx, row_data, row_idx + 1)

    num_rows = len(data)
    red_fill = PatternFill("solid", fgColor=RED_FILL)
    orange_fill = PatternFill("solid", fgColor=AMBER_FILL)

    ws.conditional_formatting.add(
        f"C2:C{num_rows + 1}",
        CellIsRule(operator="equal", formula=['"P1"'], fill=red_fill)
    )
    ws.conditional_formatting.add(
        f"C2:C{num_rows + 1}",
        CellIsRule(operator="equal", formula=['"P2"'], fill=orange_fill)
    )

    _freeze_and_filter(ws, f"A1:M{num_rows + 1}")
    _auto_size_columns(ws)


# ---------------------------------------------------------------------------
# Sheet 4: UAT TRACKER
# ---------------------------------------------------------------------------

def create_uat_tracker(wb):
    ws = wb.create_sheet("UAT TRACKER")

    headers = [
        "UAT ID", "Linked Sprint Ticket", "Feature/Story", "UAT Owner",
        "Date Sent to UAT", "UAT Status", "Defects Raised", "Defect Severity",
        "Date Resolved", "UAT Sign-off By", "Sign-off Date", "PROD Ready?",
    ]
    _apply_header(ws, headers)

    data = [
        ("UAT-001", "PROJ-101", "User authentication via SSO", "Sarah Connor",
         "2026-05-29", "Passed", 1, "Minor", "2026-05-30",
         "Sarah Connor", "2026-05-30", "Yes"),
        ("UAT-002", "PROJ-103", "Payment gateway timeout handling", "Mark Davis",
         "2026-05-30", "In Testing", 2, "Major", "",
         "", "", "No"),
        ("UAT-003", "PROJ-105", "Export to PDF feature", "Sarah Connor",
         "2026-05-31", "Pending", 0, "", "",
         "", "", "No"),
        ("UAT-004", "PROJ-102", "Dashboard widget rendering", "Mark Davis",
         "2026-05-28", "Failed", 3, "Critical", "",
         "", "", "No"),
    ]

    for row_idx, row_data in enumerate(data, start=1):
        _apply_data_row(ws, row_idx, row_data, row_idx + 1)

    num_rows = len(data)
    green_fill = PatternFill("solid", fgColor=GREEN_FILL)
    red_fill = PatternFill("solid", fgColor=RED_FILL)
    amber_fill = PatternFill("solid", fgColor=AMBER_FILL)

    ws.conditional_formatting.add(
        f"F2:F{num_rows + 1}",
        CellIsRule(operator="equal", formula=['"Passed"'], fill=green_fill)
    )
    ws.conditional_formatting.add(
        f"F2:F{num_rows + 1}",
        CellIsRule(operator="equal", formula=['"Failed"'], fill=red_fill)
    )
    ws.conditional_formatting.add(
        f"F2:F{num_rows + 1}",
        CellIsRule(operator="equal", formula=['"In Testing"'], fill=amber_fill)
    )

    _freeze_and_filter(ws, f"A1:L{num_rows + 1}")
    _auto_size_columns(ws)


# ---------------------------------------------------------------------------
# Sheet 5: QA ISSUES
# ---------------------------------------------------------------------------

def create_qa_issues(wb):
    ws = wb.create_sheet("QA ISSUES")

    headers = [
        "QA Issue ID", "Linked Sprint Ticket", "Summary", "Severity",
        "Found By", "Date Found", "Assigned Dev", "Status",
        "Retest Date", "Retest Result", "Closed Date",
        "Blocks UAT?", "Blocks PROD?",
    ]
    _apply_header(ws, headers)

    data = [
        ("QA-001", "PROJ-102", "Widget tooltip overlaps chart on mobile", "High",
         "David Lee", "2026-05-26", "Bob Smith", "In Fix",
         "2026-05-29", "", "", "No", "No"),
        ("QA-002", "PROJ-103", "Payment timeout not retried after 30s", "Critical",
         "David Lee", "2026-05-27", "Carol White", "Retest",
         "2026-05-30", "Pass", "", "Yes", "Yes"),
        ("QA-003", "PROJ-101", "SSO logout token not invalidated on server", "High",
         "David Lee", "2026-05-28", "Alice Johnson", "Closed",
         "2026-05-29", "Pass", "2026-05-29", "No", "No"),
        ("QA-004", "PROJ-105", "PDF export crops footer on A4 landscape", "Medium",
         "Emma Brown", "2026-05-30", "Eve Martinez", "Open",
         "", "", "", "No", "Yes"),
    ]

    for row_idx, row_data in enumerate(data, start=1):
        _apply_data_row(ws, row_idx, row_data, row_idx + 1)

    num_rows = len(data)
    red_fill = PatternFill("solid", fgColor=RED_FILL)
    green_fill = PatternFill("solid", fgColor=GREEN_FILL)

    ws.conditional_formatting.add(
        f"H2:H{num_rows + 1}",
        CellIsRule(operator="equal", formula=['"Open"'], fill=red_fill)
    )
    ws.conditional_formatting.add(
        f"H2:H{num_rows + 1}",
        CellIsRule(operator="equal", formula=['"Closed"'], fill=green_fill)
    )

    _freeze_and_filter(ws, f"A1:M{num_rows + 1}")
    _auto_size_columns(ws)


# ---------------------------------------------------------------------------
# Sheet 6: PROD RELEASE TRACEABILITY
# ---------------------------------------------------------------------------

def create_release_traceability(wb):
    ws = wb.create_sheet("PROD RELEASE TRACEABILITY")

    headers = [
        "Release ID", "Sprint Ticket ID", "Feature Summary",
        "Dev Done Date", "QA Status", "QA Completed Date",
        "UAT Status", "UAT Sign-off Date", "UAT Sign-off By",
        "PROD Deploy Date", "PROD Deploy Status", "Deployed By",
        "Post-Deploy Verified?", "Rollback Plan?", "Notes",
    ]
    _apply_header(ws, headers)

    data = [
        ("REL-2026-05-31", "PROJ-101", "User authentication via SSO",
         "2026-05-28", "Pass", "2026-05-29",
         "Passed", "2026-05-30", "Sarah Connor",
         "2026-05-31", "Deployed", "DevOps Bot",
         "Yes", "Yes", "Smooth deploy; no issues observed"),
        ("REL-2026-05-31", "PROJ-102", "Dashboard widget rendering fix",
         "", "In Progress", "",
         "Pending", "", "",
         "2026-05-31", "Scheduled", "DevOps Bot",
         "No", "Yes", "Blocked – QA not yet complete"),
        ("REL-2026-06-05", "PROJ-103", "Payment gateway timeout handling",
         "", "In Progress", "",
         "In Testing", "", "Mark Davis",
         "2026-06-05", "Scheduled", "DevOps Bot",
         "No", "Yes", "P1 incident INC-001 linked; high priority release"),
    ]

    for row_idx, row_data in enumerate(data, start=1):
        _apply_data_row(ws, row_idx, row_data, row_idx + 1)

    num_rows = len(data)
    green_fill = PatternFill("solid", fgColor=GREEN_FILL)
    red_fill = PatternFill("solid", fgColor=RED_FILL)
    amber_fill = PatternFill("solid", fgColor=AMBER_FILL)

    ws.conditional_formatting.add(
        f"K2:K{num_rows + 1}",
        CellIsRule(operator="equal", formula=['"Deployed"'], fill=green_fill)
    )
    ws.conditional_formatting.add(
        f"K2:K{num_rows + 1}",
        CellIsRule(operator="equal", formula=['"Rolled Back"'], fill=red_fill)
    )
    ws.conditional_formatting.add(
        f"K2:K{num_rows + 1}",
        CellIsRule(operator="equal", formula=['"Scheduled"'], fill=amber_fill)
    )

    _freeze_and_filter(ws, f"A1:O{num_rows + 1}")
    _auto_size_columns(ws)


# ---------------------------------------------------------------------------
# Sheet 7: DAILY STANDUP LOG
# ---------------------------------------------------------------------------

def create_standup_log(wb):
    ws = wb.create_sheet("DAILY STANDUP LOG")

    headers = [
        "Date", "Team", "Member", "Yesterday", "Today",
        "Blockers", "Action Owner", "Action Due", "Resolved?",
    ]
    _apply_header(ws, headers)

    data = [
        ("2026-05-29", "Backend", "Alice Johnson",
         "Completed SSO integration and unit tests",
         "Code review for PROJ-103; start payment retry logic",
         "None", "", "", "N/A"),
        ("2026-05-29", "Frontend", "Bob Smith",
         "Worked on dashboard widget layout",
         "Fix tooltip overlap issue (QA-001); design review",
         "Figma assets not delivered by design team",
         "Alice Johnson", "2026-05-30", "No"),
        ("2026-05-30", "QA", "David Lee",
         "Raised QA-001 and QA-002; regression run on auth module",
         "Retest PROJ-103 payment timeout after Carol's fix",
         "None", "", "", "N/A"),
        ("2026-05-30", "Backend", "Carol White",
         "Investigated P1 INC-001 DB connection pool exhaustion",
         "Apply index migration to PROD orders table; update PROJ-110",
         "Vendor API rate limit — escalated, waiting for response",
         "Carol White", "2026-05-31", "No"),
        ("2026-05-31", "Frontend", "Eve Martinez",
         "Reviewed PDF export requirements (PROJ-105)",
         "Begin PDF layout implementation; address QA-004 footer issue",
         "None", "", "", "N/A"),
    ]

    for row_idx, row_data in enumerate(data, start=1):
        _apply_data_row(ws, row_idx, row_data, row_idx + 1)

    num_rows = len(data)
    _freeze_and_filter(ws, f"A1:I{num_rows + 1}")
    _auto_size_columns(ws)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def generate_workbook(output_path="daily_project_tracker.xlsx"):
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    sprint_sheet_title = "SPRINT TRACKER"

    create_dashboard(wb, sprint_sheet_title)
    create_sprint_tracker(wb)
    create_production_issues(wb)
    create_uat_tracker(wb)
    create_qa_issues(wb)
    create_release_traceability(wb)
    create_standup_log(wb)

    # Apply tab colours
    for ws, colour in zip(wb.worksheets, TAB_COLOURS):
        ws.sheet_properties.tabColor = colour

    wb.save(output_path)
    print(f"✅ Tracker saved to: {output_path}")


if __name__ == "__main__":
    generate_workbook()
